"""Gera bolao/Bolao_Copa2026.xlsx. Rodar uma vez (e re-rodar se Config default mudar pós-backtest).
As fórmulas são escritas em en-US canônico (openpyxl); o Excel pt-BR exibe traduzido.
Nota: POISSON.DIST é função pós-2007 — gravada como _xlfn.POISSON.DIST no XML (exigência do
formato xlsx; o Excel exibe POISSON.DIST normalmente; sem o prefixo daria #NAME?)."""
import openpyxl

from bolao.lambdas import p_casa_vence
from bolao.rubrica import pontos
from bolao.grade import MAX_GOLS
import math

N = MAX_GOLS + 1          # 7
PLACARES = [(h, a) for h in range(N) for a in range(N)]  # idx = h*7+a
ROTULOS = [f"{h}x{a}" for h, a in PLACARES]

def col(idx1):
    return openpyxl.utils.get_column_letter(idx1)

def build_config(wb):
    ws = wb.create_sheet("Config")
    # θ=0: o shade -0.05 foi calibrado em fechamento Pinnacle e empilha na MESMA direção
    # do viés conhecido do devig proporcional em casas soft (red team 12/06). ρ=1.10:
    # 1.18 não tem suporte fora de clubes e o input soft já chega com empate inflado.
    # δ=0.05 em unidades de fase 1x (ver grade.decisao).
    linhas = [("Parâmetro", "Valor"),
              ("theta_shade", 0.0), ("rho_empate", 1.10), ("delta_desempate", 0.05),
              ("w_por_flag", 0.05), ("cap_excecao_pct_lambda", 0.10), ("tripwire_gols", 0.4)]
    for i, (a, b) in enumerate(linhas, start=1):
        ws.cell(i, 1, a); ws.cell(i, 2, b)

def build_tabelas(wb):
    ws = wb.create_sheet("Tabelas")
    ws.cell(1, 1, "lambda_total"); ws.cell(1, 2, "p_over_2.5")
    for i in range(581):  # 0.20..6.00 passo 0.01, sem acúmulo de erro float
        lt = round(0.20 + 0.01 * i, 2)
        p_over = 1 - math.exp(-lt) * (1 + lt + lt * lt / 2)
        ws.cell(2 + i, 1, lt); ws.cell(2 + i, 2, p_over)
    s_vals = [round(-3.0 + 0.05 * j, 2) for j in range(121)]
    lt_vals = [round(0.60 + 0.05 * i, 2) for i in range(109)]
    for j, s in enumerate(s_vals):
        ws.cell(600, 2 + j, s)
    for i, lt in enumerate(lt_vals):
        ws.cell(601 + i, 1, lt)
        for j, s in enumerate(s_vals):
            if abs(s) < lt:
                v = p_casa_vence(lt, s)
            else:
                v = 0.0 if s < 0 else 1.0   # sentinelas monotônicas: MATCH clampa no S válido extremo
            ws.cell(601 + i, 2 + j, v)

def build_pontos(wb):
    ws = wb.create_sheet("Pontos")
    ws.cell(1, 1, "resultado\\palpite")
    for jdx, rot in enumerate(ROTULOS):
        ws.cell(1, 2 + jdx, rot)
    for idx, rot in enumerate(ROTULOS):
        ws.cell(2 + idx, 1, rot)
        for jdx in range(len(ROTULOS)):
            ws.cell(2 + idx, 2 + jdx, pontos(PLACARES[jdx], PLACARES[idx]))

def build_jogo(wb):
    ws = wb.create_sheet("Jogo")
    rotin = [("Odd Over 2.5", 2), ("Odd Under 2.5", 3), ("Odd Casa", 4), ("Odd Empate", 5),
             ("Odd Fora", 6), ("Fase (1 grupo/2 mata/3 final)", 7),
             ("Δλ exceção CASA (POLITICA.md)", 8), ("Δλ exceção FORA", 9)]
    for nome, r in rotin:
        ws.cell(r, 1, nome)
    ws.cell(7, 2, 1); ws.cell(8, 2, 0); ws.cell(9, 2, 0)
    flags = ["Flag estreia", "Flag dead rubber", "Flag clima extremo",
             "Flag risco rotação", "Flag dados ralos"]
    for i, f in enumerate(flags):
        ws.cell(2 + i, 3, f); ws.cell(2 + i, 4, 0)
    ws.cell(2, 5, "Modal pelotão (ex 2x1)"); ws.cell(2, 6, "")
    ws.cell(3, 5, "Estado (LIDER/PELOTAO/CACADOR)"); ws.cell(3, 6, "PELOTAO")
    ws.cell(5, 5, "S_elo (aba Elo, opcional)"); ws.cell(5, 6, "")
    calc = [
        (11, "p_over", "=(1/B2)/((1/B2)+(1/B3))"),
        (12, "p_casa", "=(1/B4)/((1/B4)+(1/B5)+(1/B6))"),
        (13, "lambda_total", "=INDEX(Tabelas!$A$2:$A$582,MATCH(B11,Tabelas!$B$2:$B$582,1))"),
        (14, "linha_T2", "=MATCH(B13,Tabelas!$A$601:$A$709,1)"),
        (15, "S (supremacia)",
         "=INDEX(Tabelas!$B$600:$DR$600,MATCH(B12,INDEX(Tabelas!$B$601:$DR$709,B14,0),1))"),
        (16, "lambda_casa_mercado", "=(B13+B15)/2"),
        (17, "lambda_fora_mercado", "=(B13-B15)/2"),
        (18, "lambda_casa_final",
         "=B16*(1+Config!$B$2)+MAX(-Config!$B$6*B16,MIN(Config!$B$6*B16,B8))"),
        (19, "lambda_fora_final",
         "=B17*(1+Config!$B$2)+MAX(-Config!$B$6*B17,MIN(Config!$B$6*B17,B9))"),
        (20, "w achatamento", "=MIN(0.25,Config!$B$5*SUM(D2:D6))"),
    ]
    for r, nome, f in calc:
        ws.cell(r, 1, nome); ws.cell(r, 2, f)
    for h in range(N):
        ws.cell(2 + h, 7, h)
        ws.cell(1, 8 + h, h)
    for h in range(N):
        for a in range(N):
            rho = "*Config!$B$3" if (h == a and h <= 1) else ""
            ws.cell(2 + h, 8 + a,
                    f"=_xlfn.POISSON.DIST($G{2+h},$B$18,FALSE)"
                    f"*_xlfn.POISSON.DIST({col(8+a)}$1,$B$19,FALSE){rho}")
    ws.cell(10, 15, "soma grid"); ws.cell(10, 16, "=SUM(H2:N8)")
    cab = ["placar", "ph", "pa", "P", "EV", "P_exato", "P_invertido"]
    for j, c in enumerate(cab):
        ws.cell(12, 7 + j, c)
    for idx, (h, a) in enumerate(PLACARES):
        r = 13 + idx
        ws.cell(r, 7, ROTULOS[idx]); ws.cell(r, 8, h); ws.cell(r, 9, a)
        ws.cell(r, 10, f"=((1-$B$20)*INDEX($H$2:$N$8,H{r}+1,I{r}+1)/$P$10)+$B$20/49")
        ws.cell(r, 11, f"=SUMPRODUCT(INDEX(Pontos!$B$2:$AX$50,0,{idx + 1}),$J$13:$J$61)*$B$7")
        ws.cell(r, 12, f"=J{r}")
        ws.cell(r, 13, f"=IF(H{r}=I{r},0,INDEX($J$13:$J$61,I{r}*7+H{r}+1))")
    res = [(2, "EV 1º", "=LARGE($K$13:$K$61,1)", "=INDEX($G$13:$G$61,MATCH(Q2,$K$13:$K$61,0))"),
           (3, "EV 2º", "=LARGE($K$13:$K$61,2)", "=INDEX($G$13:$G$61,MATCH(Q3,$K$13:$K$61,0))"),
           (4, "EV 3º", "=LARGE($K$13:$K$61,3)", "=INDEX($G$13:$G$61,MATCH(Q4,$K$13:$K$61,0))")]
    for r, nome, fq, fr in res:
        ws.cell(r, 16, nome); ws.cell(r, 17, fq); ws.cell(r, 18, fr)
    ws.cell(5, 16, "margem"); ws.cell(5, 17, "=Q2-Q3")
    ws.cell(6, 16, "decisão")
    ws.cell(6, 17, '=IF(Q5<=Config!$B$4,"FRONTEIRA: aplicar POLITICA.md vs modal '
                   'F2","OK: cravar o 1º")')
    ws.cell(7, 16, "tripwire")
    ws.cell(7, 17, '=IF($F$5="","sem elo",IF(ABS(B15-$F$5)>Config!$B$7,'
                   '"⚠ CONFERIR LINHA DIGITADA","ok"))')

def build_validacao(wb):
    ws = wb.create_sheet("Validação")
    ws.cell(1, 1, "Caso"); ws.cell(1, 2, "palpite"); ws.cell(1, 3, "resultado")
    ws.cell(1, 4, "esperado"); ws.cell(1, 5, "matriz Pontos"); ws.cell(1, 6, "dif")
    from bolao.tests.test_rubrica import CASOS
    for i, (ph, pa, rh, ra, esp, desc) in enumerate(CASOS):
        r = 2 + i
        jdx = ph * 7 + pa
        idx = rh * 7 + ra
        ws.cell(r, 1, desc); ws.cell(r, 2, f"{ph}x{pa}"); ws.cell(r, 3, f"{rh}x{ra}")
        ws.cell(r, 4, esp)
        ws.cell(r, 5, f"=INDEX(Pontos!$B$2:$AX$50,{idx + 1},{jdx + 1})")
        ws.cell(r, 6, f"=E{r}-D{r}")
    # F26 é contrato de layout (valida_planilha.py lê F26); o range acompanha len(CASOS)
    last_row = 1 + len(CASOS)
    assert last_row <= 24, "CASOS cresceu além do layout: ajustar Validação e valida_planilha.py"
    ws.cell(26, 1, "SOMA |dif| (tem que ser 0):")
    ws.cell(26, 6, f"=SUMPRODUCT(ABS(F2:F{last_row}))")

def build_politica_e_log(wb):
    ws = wb.create_sheet("Política")
    ws.cell(1, 1, "Conteúdo oficial em bolao/POLITICA.md — congelado 10/06/2026."
                  " LIDER=colar no modal | PELOTAO=EV+desempate divergente |"
                  " CACADOR=variância nos jogos com multiplicador."
                  " Exceções de λ: SÓ as 3 da lista fechada, teto ±10%.")
    ws = wb.create_sheet("Log")
    cab = ["data", "jogo", "fase", "odds (5)", "λT", "S", "λcasa", "λfora", "flags",
           "exceção?", "palpite", "EV", "margem", "modal pelotão", "estado", "racional",
           "resultado", "pontos"]
    for j, c in enumerate(cab, start=1):
        ws.cell(1, j, c)

def main():
    import os
    if os.path.exists("bolao/Bolao_Copa2026.xlsx"):
        print("AVISO: regenerar APAGA dados digitados nas abas Jogo e Log "
              "(copie o Log antes, se houver histórico).")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_config(wb)
    build_tabelas(wb)
    build_pontos(wb)
    build_jogo(wb)
    build_validacao(wb)
    build_politica_e_log(wb)
    wb.save("bolao/Bolao_Copa2026.xlsx")
    print("OK: bolao/Bolao_Copa2026.xlsx gerada.")

if __name__ == "__main__":
    main()
