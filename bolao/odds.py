"""Captura de odds + motor num comando só (zero token, ~10s).

uso: python -m bolao.odds <time1> <time2> [--fase 1|2|3] [--flags N]

Ex.: python -m bolao.odds coreia tcheca
     python -m bolao.odds brasil marrocos --flags 1

Fonte: SportyTrader (mediana das casas listadas; páginas renderizadas via r.jina.ai
porque as tabelas são JS). Parâmetros θ/ρ/δ/w lidos da Config da planilha (fonte única).
Times em PT ou EN; a descoberta casa substrings com o slug da URL do jogo."""
import re
import statistics
import sys
import unicodedata
import warnings

warnings.filterwarnings("ignore")
import requests

if sys.stdout is not None:  # pythonw (Agendador) roda sem stdout
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
COMPETICAO = "https://www.sportytrader.com/en/odds/football/world/world-cup-1811/"
JINA = "https://r.jina.ai/"

# nome comum (PT/EN, sem acento) -> fragmento do slug em inglês
SLUGS = {
    "mexico": "mexico", "africa do sul": "south-africa", "coreia": "south-korea",
    "tcheca": "czech", "tchequia": "czech", "brasil": "brazil", "marrocos": "morocco",
    "haiti": "haiti", "argentina": "argentina", "argelia": "algeria",
    "austria": "austria", "jordania": "jordan", "espanha": "spain", "franca": "france",
    "inglaterra": "england", "portugal": "portugal", "alemanha": "germany",
    "holanda": "netherlands", "italia": "italy", "belgica": "belgium", "ira": "iran",
    "egito": "egypt", "croacia": "croatia", "uruguai": "uruguay", "colombia": "colombia",
    "equador": "ecuador", "eua": "usa", "estados unidos": "usa", "canada": "canada",
    "japao": "japan", "senegal": "senegal", "gana": "ghana", "suica": "switzerland",
    "dinamarca": "denmark", "noruega": "norway", "australia": "australia",
    "turquia": "turkey", "catar": "qatar", "bosnia": "bosnia", "polonia": "poland",
    "escocia": "scotland", "tunisia": "tunisia", "costa do marfim": "ivory-coast",
    "nigeria": "nigeria", "paraguai": "paraguay", "panama": "panama",
    "costa rica": "costa-rica", "honduras": "honduras", "jamaica": "jamaica",
    "nova zelandia": "new-zealand", "uzbequistao": "uzbekistan", "cabo verde": "cape-verde",
    "africa": "south-africa", "arabia": "saudi-arabia", "curacao": "curacao",
}


def _ascii(s):
    return unicodedata.normalize("NFKD", s.lower()).encode("ascii", "ignore").decode()


def _get(url, timeout=30):
    r = requests.get(url, timeout=timeout, verify=False, headers=UA)
    r.raise_for_status()
    return r.text


def achar_jogo(t1, t2):
    """Acha a URL do jogo na página da competição casando os dois times no slug."""
    f1 = SLUGS.get(_ascii(t1), _ascii(t1).replace(" ", "-"))
    f2 = SLUGS.get(_ascii(t2), _ascii(t2).replace(" ", "-"))
    try:
        html = _get(COMPETICAO, timeout=30)
        links = sorted(set(re.findall(r'href="(/en/odds/[a-z0-9-]+-\d{6,}/)"', html)))
    except Exception:
        # SportyTrader às vezes 403 no fetch direto (Cloudflare/proxy corporativo);
        # via jina (server-side) contorna. No markdown os links vêm como URL completa.
        html = _get(JINA + COMPETICAO, timeout=60)
        links = sorted(set(re.findall(
            r"https://www\.sportytrader\.com(/en/odds/[a-z0-9-]+-\d{6,}/)", html)))
    cand = [u for u in links if f1 in u and f2 in u]
    if not cand:
        cand = [u for u in links if f1 in u or f2 in u]
        raise SystemExit(f"jogo '{t1} x {t2}' não encontrado (fragmentos: {f1}, {f2}).\n"
                         f"Candidatos parciais: {cand[:8]}")
    if len(cand) > 1:
        print(f"AVISO: {len(cand)} jogos casam — usando o primeiro: {cand}")
    return "https://www.sportytrader.com" + cand[0]


# A render do jina oscila entre DOIS layouts (medido 26/06, mesmo instante, jogos
# diferentes): (A) "flat" — texto achatado sem logos, cada cota numa linha, com a
# célula de bônus (€/$/£ ou "See the offer") aparecendo de forma INCONSTANTE (vários
# jogos sem nenhuma); (B) "rich" — markdown com logo das casas, cada casa numa linha
# "100x45/<casa>.webp) o1 o2 [o3]](...". O parser aguenta os dois e NÃO depende da
# célula de bônus como terminador (o que matava o flat sem bônus). Cada mercado é uma
# seção que vai do seu título ao título do PRÓXIMO mercado conhecido — vale nos dois.
MERCADOS = (
    "Full Time Result", "Half Time Result", "Half Time/Full Time",
    "Under/Over 0.5 Goals", "Under/Over 1.5 Goals", "Under/Over 2.5 Goals",
    "Under/Over 3.5 Goals", "Under/Over 4.5 Goals", "Under/Over 5.5 Goals",
    "Both Teams To Score", "Double Chance", "Draw No Bet", "Odd/Even",
    "Correct Score", "Highest Scoring Half", "Total Corners", "Handicap",
)
_RICH = re.compile(r"100x45/[^)]+?\.webp\)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)"
                   r"(?:\s+(\d+(?:\.\d+)?))?")  # 2 ou 3 cotas inline (OU / 1X2)
_NUM = re.compile(r"\d+(?:\.\d+)?")


def _secao(txt, titulo):
    """Recorta a seção de UM mercado: do seu título até o título do próximo mercado
    conhecido (o que vier antes). Independe do layout (flat ou rich)."""
    i = txt.find(titulo)
    if i < 0:
        return ""
    fim = len(txt)
    for m in MERCADOS:
        j = txt.find(m, i + len(titulo))
        if 0 < j < fim:
            fim = j
    return txt[i:fim]


def _linhas(bloco, k):
    """Cotas de uma seção como lista de tuplas de k cotas. Aguenta os dois layouts."""
    if "100x45/" in bloco:                       # layout B (rich): casa com logo
        linhas = []
        for m in _RICH.findall(bloco):
            vals = [v for v in m if v]
            if len(vals) >= k:
                linhas.append(tuple(vals[:k]))
        if len(linhas) >= 3:
            return linhas
    # layout A (flat): números puros após "BONUS UP TO", ignorando células de bônus
    # (inconstantes) e parando no texto do próximo mercado.
    j = bloco.find("BONUS UP TO")
    resto = bloco[j + len("BONUS UP TO"):] if j >= 0 else bloco
    nums = []
    for ln in resto.split("\n"):
        s = ln.strip()
        if not s:
            continue
        if _NUM.fullmatch(s):
            nums.append(s)
        elif s[0] in "€$£" or s.lower().startswith("see the offer"):
            continue                             # célula de bônus → ignora
        else:
            break                                # texto = próximo mercado → fim
    return [tuple(nums[x:x + k]) for x in range(0, (len(nums) // k) * k, k)]


def parse_render(txt):
    """Extrai as 5 odds medianas + qual de um render jina. Pura (testável sem rede)."""
    if "Full Time Result" not in txt or "Under/Over 2.5 Goals" not in txt:
        raise SystemExit("seções de odds não encontradas no render — tentar de novo "
                         "(jina instável) ou capturar manual (FONTES.md)")
    m1x2 = _linhas(_secao(txt, "Full Time Result"), 3)
    mou = _linhas(_secao(txt, "Under/Over 2.5 Goals"), 2)
    if len(m1x2) < 3 or len(mou) < 3:
        raise SystemExit(f"poucas casas parseadas (1X2={len(m1x2)}, OU={len(mou)}) — "
                         "layout mudou? conferir manual (FONTES.md)")
    casa = statistics.median(float(r[0]) for r in m1x2)
    emp = statistics.median(float(r[1]) for r in m1x2)
    fora = statistics.median(float(r[2]) for r in m1x2)
    under = statistics.median(float(r[0]) for r in mou)  # cabeçalho: -2.5 +2.5
    over = statistics.median(float(r[1]) for r in mou)
    disp = (f"1X2 n={len(m1x2)} casa[{min(float(r[0]) for r in m1x2)}-"
            f"{max(float(r[0]) for r in m1x2)}], OU n={len(mou)}")

    def _spread_max(rows, n_cols):
        """maior dispersão relativa entre casas, SÓ nas colunas de odds curtas (mediana
        <=6 = prob >=~1/6). Azarão (empate 17, fora 40) varia muito entre casas sem
        afetar o λ — incluí-lo dava falso positivo de 'contaminação' em jogo extremo."""
        s = 0.0
        for i in range(n_cols):
            vals = [float(r[i]) for r in rows]
            med = statistics.median(vals)
            if med <= 6.0:
                s = max(s, (max(vals) - min(vals)) / med)
        return s
    qual = {"n_1x2": len(m1x2), "n_ou": len(mou),
            "spread_1x2": _spread_max(m1x2, 3), "spread_ou": _spread_max(mou, 2),
            "overround_1x2": 1 / casa + 1 / emp + 1 / fora,
            "overround_ou": 1 / under + 1 / over}
    return {"over": over, "under": under, "casa": casa, "empate": emp, "fora": fora}, disp, qual


def capturar(url_jogo):
    """Renderiza a página do jogo e devolve as 5 odds medianas + nº de casas."""
    return parse_render(_get(JINA + url_jogo, timeout=90))


def captura_ok(qual):
    """A 'conferida' humana da captura, em código: casas suficientes, casas concordando
    entre si, e overround na faixa normal de bookmaker (margem 0.5%-18%). Linha de outro
    mercado vazando pras medianas quebra pelo menos um destes três."""
    checks = {
        "n_casas": qual["n_1x2"] >= 5 and qual["n_ou"] >= 5,
        # contaminação real (linha de OUTRO mercado vazando) dá spread >1; dispersão
        # genuína de linha cedo (T-9h) chega a ~0.35 — medido em 12/06 (USA x Paraguay)
        "dispersao": qual["spread_1x2"] <= 0.5 and qual["spread_ou"] <= 0.5,
        "overround": (1.005 <= qual["overround_1x2"] <= 1.18
                      and 1.005 <= qual["overround_ou"] <= 1.15),
    }
    return all(checks.values()), checks


def config_da_planilha():
    """θ/ρ/δ/w_por_flag da aba Config (fonte única). Fallback p/ os defaults congelados
    (POLITICA: θ=0/ρ=1.10/δ=0.05/w=0.05) se o xlsx/openpyxl não estiver presente — ex.:
    rodando na nuvem (GitHub Actions). Params estão congelados, então o default é fiel."""
    import os
    fp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Bolao_Copa2026.xlsx")
    try:
        import openpyxl
        ws = openpyxl.load_workbook(fp, read_only=True)["Config"]
        v = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, 8)}
        return (v["theta_shade"], v["rho_empate"], v["delta_desempate"], v["w_por_flag"])
    except Exception:
        return (0.0, 1.10, 0.05, 0.05)


def estado_ranking():
    """LIDER/PELOTAO/CACADOR de dados/auto_config.json (editar quando o ranking mudar)."""
    import json
    import os
    fp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados", "auto_config.json")
    try:
        with open(fp, encoding="utf-8") as f:
            e = json.load(f).get("estado", "PELOTAO").upper()
            return e if e in ("LIDER", "PELOTAO", "CACADOR") else "PELOTAO"
    except (OSError, ValueError):
        return "PELOTAO"


def analisar(odds, fase=1, nflags=0, estado=None):
    """Roda o motor sobre as 5 odds, aplica a POLITICA (resolução de fronteira em
    grade.escolher_palpite) e devolve (relatorio_texto, info) — usado pelo CLI e
    pelo auto_palpite. info traz palpite final, regra aplicada, λT, S e os p's."""
    from bolao.grade import grade_probabilidades, achatar, escolher_palpite, ev_palpites
    from bolao.lambdas import (prob_implicita_2, prob_implicita_1x2, total_de_ou25,
                               supremacia_de_1x2)
    theta, rho, delta, w_flag = config_da_planilha()
    estado = estado or estado_ranking()
    p_over = prob_implicita_2(odds["over"], odds["under"])
    p_casa, p_emp, p_fora = prob_implicita_1x2(odds["casa"], odds["empate"], odds["fora"])
    # mesmos clamps do treino (backtest): favorito extremo estourava o brentq a T-50
    p_over = min(max(p_over, 0.02), 0.98)
    p_casa_c = min(max(p_casa, 0.02), 0.96)
    lt = total_de_ou25(p_over)
    linhas = []
    try:
        s = supremacia_de_1x2(lt, p_casa_c)
        s_ok = True
    except ValueError:
        # p_casa maior do que o total de gols permite: combo incoerente (captura
        # contaminada?) — clampa S no máximo solúvel em vez de morrer sem palpite
        s = lt - 0.02
        s_ok = False
        linhas.append(f"⚠ p_casa={p_casa_c:.3f} incompatível com λT={lt:.2f} — "
                      "S clampado no máximo; CONFERIR captura")
    if lt < 1.0 or lt > 4.5:
        linhas.append(f"⚠ λT={lt:.2f} fora do range plausível — conferir captura (colunas O/U?)")
    lh, la = (lt + s) / 2 * (1 + theta), (lt - s) / 2 * (1 + theta)
    w = min(0.25, w_flag * nflags)
    g = achatar(grade_probabilidades(lh, la, rho=rho), w=w)
    r = escolher_palpite(g, multiplicador=fase, delta=delta, estado=estado,
                         p_extremo=max(p_casa, p_fora))
    d = r["decisao"]
    linhas.append(f"params Config: θ={theta} ρ={rho} δ={delta} | fase={fase} "
                  f"flags={nflags} (w={w}) | estado={estado}")
    linhas.append(f"λT={lt:.3f} S={s:.3f} → λ ajustado=({lh:.3f}, {la:.3f}) | "
                  f"p=({p_casa:.3f}/{p_emp:.3f}/{p_fora:.3f})")
    modal = r["modal"]
    linhas.append(f"célula modal da grade (pelotão previsto): {modal[0]}x{modal[1]} "
                  f"P={g[modal[0]][modal[1]]:.3f}")
    linhas.append("TOP-5 por EV:")
    for (ph, pa), ev in ev_palpites(g, multiplicador=fase)[:5]:
        pinv = 0.0 if ph == pa else g[pa][ph]
        linhas.append(f"  {ph}x{pa}: EV={ev:.4f}  P_exato={g[ph][pa]:.4f}  P_inv={pinv:.4f}")
    ph, pa = r["palpite"]
    linhas.append(f"margem={d['margem']:.4f} | {r['regra']}")
    linhas.append(f"PALPITE FINAL: {ph}x{pa}")
    info = {"palpite": r["palpite"], "regra": r["regra"], "fronteira": d["fronteira"],
            "modal": modal, "lt": lt, "s": s, "p_casa": p_casa, "p_emp": p_emp,
            "p_fora": p_fora, "estado": estado,
            "lt_ok": bool(1.0 <= lt <= 4.5) and s_ok}  # False => ⚠ CONFERIR no assunto
    return "\n".join(linhas), info


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    fase = int(sys.argv[sys.argv.index("--fase") + 1]) if "--fase" in sys.argv else 1
    nflags = int(sys.argv[sys.argv.index("--flags") + 1]) if "--flags" in sys.argv else 0
    if len(args) != 2:
        print(__doc__)
        raise SystemExit(1)
    t1, t2 = args
    url = achar_jogo(t1, t2)
    print(f"jogo: {url}")
    odds, disp, qual = capturar(url)
    cap_ok, checks = captura_ok(qual)
    print(f"odds (medianas): Over2.5={odds['over']:.2f}  Under2.5={odds['under']:.2f}  "
          f"Casa={odds['casa']:.2f}  Empate={odds['empate']:.2f}  "
          f"Fora={odds['fora']:.2f}   [{disp}]")
    print("captura auto-verificada: " + ("OK" if cap_ok else f"⚠ SUSPEITA {checks}"))
    relatorio, info = analisar(odds, fase=fase, nflags=nflags)
    print(relatorio)
    print(f"tripwire: python -m bolao.elo_s \"{t1}\" \"{t2}\" {info['lt']:.2f}  (+ --neutro se aplicável)")


if __name__ == "__main__":
    main()
